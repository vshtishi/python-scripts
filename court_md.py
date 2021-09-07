import openpyxl
import os
path = os.environ['path']

theFile = openpyxl.load_workbook(path)
current_sheet = theFile['TypesStructure MetaData']
count = 0


T = []
for r in range(1, 26):
    T.insert(count, [current_sheet.cell(row=r, column=3).value, current_sheet.cell(row=r, column=5).value, current_sheet.cell(row=r, column=6).value, current_sheet.cell(row=r, column=7).value])
    T.insert(count, [current_sheet.cell(row=r, column=3).value, current_sheet.cell(row=r, column=7).value])


for row in T:
     print(row, end="")
     print(',')





