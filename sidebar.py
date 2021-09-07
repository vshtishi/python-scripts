import openpyxl
import os

theFile = openpyxl.load_workbook('/home/user/Documents/court.xlsx')
current_sheet = theFile['Sheet1']
count = 0
path  = os.environ['path']

with os.scandir(path) as entries:
     for entry in entries:
         print(entry.name)
files = []
fname = []
i = 0
for root, d_names, f_names in os.walk(path):
    for f in f_names:
        fname.append(root)
        files.append(f)

for name in fname:
    file_path = name.replace(path, '')
    print(file_path)
    count += 1
    current_sheet.cell(row=count, column=1).value = file_path
    current_sheet.cell(row=count, column=2).value = '/forms/' + file_path + '/' + files[i]
    i += 1


theFile.save('/home/user/court.xlsx')
