import openpyxl
import links
import os
path  = os.environ['path']
theFile = openpyxl.load_workbook(path)
current_sheet = theFile['Sheet1']
count = 0

for r in range(2, 1786):
 cell = current_sheet.cell(row=r, column=3).value
 cell = cell.rstrip('/')

 if cell in links.meta_links:
      count+=1
      current_sheet.cell(row=r, column=5).value = 1
      current_sheet.cell(row=r, column=6).value = 1
      current_sheet.cell(row=r, column=7).value = 1



print(count)

theFile.save('/home/user/Pages.xlsx')
