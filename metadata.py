import openpyxl
import os
path = os.environ['path']

theFile = openpyxl.load_workbook(path)
current_sheet = theFile['Tier 3']
count = 0
T = []
for r in range(792, 843):
      T.insert(count, [current_sheet.cell(row=r, column=2).value, current_sheet.cell(row=r, column=5).value, '', ''])
      T.insert(count, [current_sheet.cell(row=r, column=2).value, current_sheet.cell(row=r, column=5).value, current_sheet.cell(row=r, column=6).value, current_sheet.cell(row=r, column=7).value])
      current_sheet.cell(row=r, column=5).value = 1
      current_sheet.cell(row=r, column=6).value = 1
      current_sheet.cell(row=r, column=7).value = 1



for row in T:
   print(row, end="")
   print(',')


